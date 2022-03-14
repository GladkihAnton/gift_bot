from tempfile import NamedTemporaryFile

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class OrderExcel:
    HEADERS = [
        'Получатель',
        'Повод',
        'Пол',
        'Статус заказа',
        'Дата доставки',
        'Адрес доставки',
        'Контактное лицо',
        'Комментарий',
    ]

    RECIPIENT_NAME_COLUMN = 1
    HOLIDAY_COLUMN = 2
    SEX_COLUMN = 3
    STATUS_ORDER_COLUMN = 4
    DELIVERED_AT_COLUMN = 5
    DELIVERY_ADDRESS = 6
    CONTACT_INFO = 7
    COMMENT = 8

    STATUS_ORDER_TO_COLOR = {
        'Выбран': 'FCF3D1',
        'Куплен': 'F6D979',
        'Ожидает проверки': '94B4F3',
        'Качество проверено': '90D49A',
        'Доставлен': '58A65C',
    }

    def __init__(self):
        self.wb = Workbook()
        self.ws: Worksheet = self.wb.active

        self.font_italic = Font(italic=True)
        self.font_bold = Font(bold=True)
        self.horizontal = Alignment(horizontal='center')

    def fill_orders(self, orders) -> None:
        self._prepare_headers()
        self._fill_data(orders)
        self._update_width()

    def save(self) -> str:
        with NamedTemporaryFile(delete=False) as tmp:
            self.wb.save(tmp.name)
            return tmp.name

    def _prepare_headers(self):
        for index, header in enumerate(self.HEADERS, start=1):
            cell = self.ws.cell(1, index, header)
            cell.font = self.font_italic
            cell.alignment = self.horizontal

    def _fill_data(self, rows):
        for row, row_data in enumerate(rows, start=2):
            for column, value in enumerate(row_data, start=1):
                cell = self.ws.cell(row, column, value)
                cell.alignment = self.horizontal

                if column == self.RECIPIENT_NAME_COLUMN:
                    cell.font = self.font_bold

                if column == self.SEX_COLUMN:
                    if value.lower().startswith('муж'):
                        cell.fill = PatternFill('solid', start_color='CCFDEB')
                    if value.lower().startswith('жен'):
                        cell.fill = PatternFill('solid', start_color='E6D2DB')

                if column == self.STATUS_ORDER_COLUMN:
                    cell.fill = PatternFill(
                        'solid', start_color=self.STATUS_ORDER_TO_COLOR[value]
                    )
                    cell.alignment = self.horizontal
                    cell.font = self.font_bold

    def _update_width(self):
        dims = {}
        for row in self.ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value)))
                    )
        for col, value in dims.items():
            self.ws.column_dimensions[col].width = (value + 2) * 1.2
